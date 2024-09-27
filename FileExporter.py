from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from DataFormatter import DataFormatter

class FileExporter:
    @staticmethod
    def export_date(current_sheet, data: list):
        FileExporter.add_dates(data, current_sheet)

    @staticmethod
    def add_dates(dates_data: list, worksheet: Worksheet):
        week = dates_data[0]
        dates = dates_data[1:]

        worksheet["A1"] = week

        for index, value in enumerate(dates, start=4):
            column_letter = get_column_letter(index)  # Get the letter of the column (A, B, C, ...)
            worksheet[f"{column_letter}1"] = value  # Place the value in row 1, column A, B, C, ...

    @staticmethod
    def add_employee_work_time(schedule: list, worksheet: Worksheet, count: int):
        column = 4 + (count - 1) * 2

        for index, value in enumerate(schedule, start=2):
            if value:
                column_letter = get_column_letter(index)  # Get the letter of the column (A, B, C, ...)
                worksheet[f"{column_letter}{column}"] = value  # Place the value in row 1, column A, B, C, ...

    @staticmethod
    def create_work_sheet(sheet_name: str, workbook: Workbook):
        template = workbook["Mall"]
        new_worksheet = workbook.copy_worksheet(template)
        new_worksheet.title = sheet_name


    @staticmethod
    def save_file(path: str, workbook: Workbook):
        workbook.save(path)
        workbook.close()

    @staticmethod
    def add_care_unit_name(care_unit_not_formatted: str, worksheet: Worksheet):
        care_unit = DataFormatter.get_care_unit_name(care_unit_not_formatted)
        worksheet["B2"] = care_unit

