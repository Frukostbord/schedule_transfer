import openpyxl
from FileExporter import FileExporter
from FileProcessing import FileProcessing
from FileReader import FileReader
from TransferStages import TransferStage
from openpyxl.worksheet.worksheet import Worksheet


class FileTransfer:
    """ Main class for handling everything
        Uses the three enums under TransferStages to perform different tasks
        1. CHECK_FILES:
            1a. Checks files and if the paths are correct
        2. PROCESS_DATA:
            2a. Processes and formats all the data, so it can easily be put in the Excel
        3. EXPORT_DATA:
            3a. Takes the formatted data to the Excel and sets it in the right cells

        Lastly, it saves the file and the program terminates
     """

    def __init__(self, path_csv_excel: str, path_excel_out: str):
        self.path_csv_excel = path_csv_excel
        self.path_excel_out = path_excel_out
        self.cleaned_data = []
        self.workbook_export = openpyxl.Workbook()
        self.start_transfer()

    def start_transfer(self):
        """ Initiates the file transfer process through various stages. """
        current_stage = TransferStage.CHECK_FILES

        while True:  # Loop until export stage is complete
            self.execute_stage(current_stage)

            if current_stage == TransferStage.EXPORT_DATA:
                print("Transfer successful!")
                break  # Exit after export data is complete

            current_stage = TransferStage(current_stage.value + 1)  # Move to the next stage

    def execute_stage(self, stage: TransferStage):
        """ Executes the logic for the current transfer stage. """
        if stage == TransferStage.CHECK_FILES:
            self.check_files()
        elif stage == TransferStage.PROCESS_DATA:
            self.process_data()
        elif stage == TransferStage.EXPORT_DATA:
            self.export_data()

    def check_files(self):
        """ Method using FileReader class to check that the files are correct"""
        try:
            # Checks to see if the files can be open properly
            FileReader.check_files(self.path_csv_excel, self.path_excel_out)
            self.set_workbook()
        except Exception as e:
            print(f"Something went wrong with the file pathways, please look over the error: {e}")

    def set_workbook(self):
        self.workbook_export = openpyxl.load_workbook(self.path_excel_out)

    def process_data(self):
        """ This method cleans the data and sets the workbook where the data is to be exported """
        try:
            # Cleans all the data, so it can be easily exported
            self.cleaned_data = FileProcessing.reformat_data_csv(self.path_csv_excel)

        except Exception as e:
            print(f"Something went wrong with the file processing, please look over the error: {e}")

    def export_data(self):
        try:
            self.copy_data_to_sheet()  # Copies data to sheet
            self.remove_template()  # Remove template worksheet
            self.save_exported_excel()  # Saves the edited file
        except Exception as e:
            print(f"Something went wrong with the file exporting, please look over the error: {e}")


    def copy_data_to_sheet(self):
        current_sheet = None
        counter = 1  # So we know which column to add which data

        for values in self.cleaned_data:
            # If the word "Vecka" is seen, it means that a new sheet has to be set
            if "Vecka" in values[0]:
                counter = 1
                week = values[0]

                self.create_sheet(week)  # Creates the worksheet
                current_sheet = self.workbook_export[week]  # Sets the current sheet
                self.add_dates(current_sheet, values)  # Adds the dates to the top column of the sheet
                self.add_care_unit_name(current_sheet)

            # There´s employee name and time to add to the sheet
            else:
                self.add_employee_work_times(values, current_sheet, counter)
                counter += 1

    def save_exported_excel(self):
        FileExporter.save_file(self.path_excel_out, self.workbook_export)

    def add_dates(self, current_sheet, dates: list):
        FileExporter.export_date(current_sheet, dates)

    def add_employee_work_times(self, worker_info: list, worksheet: Worksheet, count: int):
        FileExporter.add_employee_work_time(worker_info, worksheet, count)

    def create_sheet(self, week: str):
        FileExporter.create_work_sheet(week, self.workbook_export)  # Create a sheet with the week  # Creates the sheet

    def add_care_unit_name(self, current_sheet: Worksheet):
        FileExporter.add_care_unit_name(self.path_csv_excel, current_sheet)
    def remove_template(self):
        workbook = self.workbook_export
        FileProcessing.remove_work_sheet(workbook["Mall"], workbook)


