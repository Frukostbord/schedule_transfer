import openpyxl
from Model.FileReadingWriting.FileExporter import FileExporter
from Model.FileReadingWriting.FileProcessing import FileProcessing
from Model.Enum.TransferStages import TransferStage
from Model.FileReadingWriting.CreateCopy import CreateCopy
import Model.Data.Pathways as Pathways


class FileTransfer:
    """
    Main class for the actual data transfer between the .csv files and Excel files
    The class goes through the following stages:
    1. Create a copy for the new data, with the template in mind
    2. Process the data, by cleaning the .csv file and structuring it for writing it in the Excel file
    3. Write data in the Excel file
    4. Save, close and return True if the whole process went without problem
    """

    def __init__(self, csv_pathway: str):
        self.template_workbook = openpyxl.load_workbook(Pathways.TEMPLATE_WORKBOOK_PATHWAY)
        self.cleaned_data: list = []  # Cleaned data to be written in the Excel
        self.path_csv_excel: str = csv_pathway  # Raw data .csv file
        self.current_work_sheet = None  # Which worksheet we´re currently working on, with the data transfer
        self.workbook_export = None
        self.save_path: str = str()

    def start_transfer(self) -> TransferStage:
        """ Initiates the file transfer process through various stages. """
        current_stage = TransferStage.CREATE_COPY

        while True:  # Loop until export stage is complete
            check = self.execute_stage(current_stage)

            if current_stage == TransferStage.EXPORT_DATA and check:
                return TransferStage.TRANSFER_COMPLETE

            if not check:
                return current_stage

            current_stage = TransferStage.get_next_stage(current_stage)  # Move to the next stage

    def execute_stage(self, stage: TransferStage) -> bool:
        """ Executes the logic for the current transfer stage. """
        if stage == TransferStage.CREATE_COPY:
            return self.create_template_copy()
        elif stage == TransferStage.PROCESS_DATA:
            return self.process_data()
        elif stage == TransferStage.EXPORT_DATA:
            return self.export_data()

    def create_template_copy(self) -> bool:
        """
        Creates a copy from the template
        :return: True if it was successful, else False
        """

        try:
            save_path = CreateCopy.create_excel_copy(self.path_csv_excel)
            self.save_path = save_path
            self.workbook_export = openpyxl.load_workbook(self.save_path)
            return True
        except Exception as e:
            return False

    def process_data(self) -> bool:
        """ This method cleans the data and sets the workbook where the data is to be exported """
        try:
            # Cleans all the data, so it can be easily exported
            self.cleaned_data = FileProcessing.reformat_data_csv(self.path_csv_excel)
            return True

        except Exception as e:
            return False

    def export_data(self) -> bool:
        return self.export_all_data()

    def export_all_data(self) -> bool:
        """
        Ends the whole transfer by copying all the data to the Excel file,
        removing the template and saving the Excel
        :return: True if everything went well, else False
        """
        data_export_successful = (
                self.copy_data_to_sheet() and  # Copies data to sheet
                self.remove_template() and  # Remove template worksheet
                self.save_exported_excel())  # Saves the edited file

        return data_export_successful

    def copy_data_to_sheet(self) -> bool:
        """
        Copying data to the current sheet in the workbook
        :return: True if all data writing went well, else False
        """

        try:
            counter = 1  # So we know which column to add which data

            for values in self.cleaned_data:
                # If the word "Vecka" is seen, it means that a new sheet has to be set
                if "Vecka" in values[0]:
                    counter = 1
                    week = values[0]

                    self.create_sheet(week)  # Creates the worksheet
                    self.current_work_sheet = self.workbook_export[week]  # Sets the current sheet
                    self.add_data_to_sheet(values)  # Writes data to the sheet

                # There´s employee name and time to add to the sheet
                else:
                    self.add_employee_work_times(values, counter)
                    counter += 1

            return True
        except Exception as e:
            return False

    def add_data_to_sheet(self, dates: list) -> None:
        """
        Adds basic data to the sheet: week, dates and minimum staff
        """
        self.add_dates(self.current_work_sheet, dates)  # Adds the dates to the top column of the sheet
        self.add_care_unit_name()
        self.add_minimum_staff()

    def save_exported_excel(self) -> bool:
        """ Tries to save the file """
        try:
            FileExporter.save_file(self.workbook_export, self.save_path)
            return True
        except Exception as e:
            return False

    def add_dates(self, current_sheet, dates: list):
        """ Adds dates to the sheet """
        FileExporter.export_date(current_sheet, dates)

    def add_employee_work_times(self, worker_info: list, count: int):
        """Adds employees work time """
        FileExporter.add_employee_work_time(worker_info, self.current_work_sheet, count)

    def create_sheet(self, week: str):
        """ Creates a sheet with the week as its´ name """
        FileExporter.create_work_sheet(week, self.workbook_export)

    def add_care_unit_name(self):
        """ Adds care unit name to the sheet """
        FileExporter.add_care_unit_name(self.path_csv_excel, self.current_work_sheet)

    def remove_template(self) -> bool:
        """ Removes the template in the workbook, "Mall" means template """
        try:
            FileProcessing.remove_work_sheet(self.workbook_export["Mall"], self.workbook_export)
            return True
        except Exception as e:
            return False

    def add_minimum_staff(self):
        """ Adds minimum staff to the current sheet """
        FileExporter.add_minimum_staff(self.path_csv_excel, self.current_work_sheet)
