from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from Model.FileReadingWriting.DataFormatter import DataFormatter


class FileExporter:
    """
    The main class of writing data from the .csv file to the Excel file
    It uses different help classes to reformat data, then to add it according to the Excel´s design
    """
    @staticmethod
    def export_date(current_sheet, data: list) -> None:
        """
        Adds dates by calling another method
        :param current_sheet: current work sheet in Excel
        :param data: Data to be added
        """
        FileExporter.add_dates(data, current_sheet)

    @staticmethod
    def add_dates(dates_data: list, worksheet: Worksheet) -> None:
        """
        Adds dates to the current work sheet
        :param dates_data: dates to be added
        :param worksheet: Current work sheet
        """
        week = dates_data[0]
        dates = dates_data[1:]

        worksheet["A1"] = week  # Sets the current week in the top left corner

        # Sets the dates of each day in the Excel, by iterating through each column
        for index, value in enumerate(dates, start=4):
            column_letter = get_column_letter(index)  # Get the letter of the column (A, B, C, ...)
            worksheet[f"{column_letter}1"] = value  # Place the value in row 1, column A, B, C, ...

    @staticmethod
    def add_employee_work_time(schedule: list, worksheet: Worksheet, row: int) -> None:
        """
        This method iterates through all columns in a row and adds the worker´s work time
        :param schedule: schedule of the worker to be put in the current row
        :param worksheet: current worksheet
        :param row: Which row we´re currently at
        """

        # Simple arithmetic to add the worker´s work time in the right cell (both column and row)
        column = 4 + (row - 1) * 2

        for index, value in enumerate(schedule, start=2):
            if value:
                column_letter = get_column_letter(index)  # Get the letter of the column (A, B, C, ...)
                worksheet[f"{column_letter}{column}"] = value  # Place the value in row x and column y

    @staticmethod
    def create_work_sheet(sheet_name: str, workbook: Workbook) -> None:
        """
        This method copies the template in the Excel
        :param sheet_name: New sheet name
        :param workbook: Current work book in Excel
        """

        # Makes a copy of the template in the workbook with the new sheet name
        template = workbook["Mall"]
        new_worksheet = workbook.copy_worksheet(template)
        new_worksheet.title = sheet_name


    @staticmethod
    def save_file(workbook: Workbook, save_path_workbook: str) -> None:
        """
        Saves and closes the workbook
        """
        workbook.save(save_path_workbook)
        workbook.close()


    @staticmethod
    def add_care_unit_name(care_unit_not_formatted: str, worksheet: Worksheet) -> None:
        """
        Adds care unit name to the worksheet
        :param care_unit_not_formatted: Raw string, with care unit name in it
        :param worksheet: Current worksheet
        """
        care_unit = DataFormatter.get_care_unit_name(care_unit_not_formatted)
        worksheet["B2"] = care_unit

    @staticmethod
    def add_minimum_staff(care_unit_not_formatted: str, worksheet: Worksheet):
        """
        Adds minimum workers per care unit
        :param care_unit_not_formatted: Raw string with care unit in it
        :param worksheet:
        :return:
        """
        care_unit = DataFormatter.get_care_unit_name(care_unit_not_formatted)
        minimum_workers_formatted = DataFormatter.get_minimum_worker_formatted(care_unit)

        # Adds minimum workers as information beneath each work day
        for index, value in enumerate(range(6), start=3):
            if value:
                column_letter = get_column_letter(index)  # Get the letter of the column (A, B, C, ...)
                worksheet[f"{column_letter}3"] = minimum_workers_formatted  # Place the value in row 3, column x
